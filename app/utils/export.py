"""Excel export utilities for MSH Med Tour"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


def export_patients_to_excel(patients):
    """Export patient list to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Hastalar"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "ID", "Ad", "Soyad", "Telefon", "E-posta", "Doğum Tarihi",
        "Cinsiyet", "Uyruk", "Pasaport No", "Muayene Sayısı", "Kayıt Tarihi"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, patient in enumerate(patients, 2):
        # Format gender
        gender_display = "Erkek" if patient.gender == "M" else ("Kadın" if patient.gender == "F" else "-")
        
        data = [
            patient.id,
            patient.first_name,
            patient.last_name,
            patient.phone or "-",
            patient.email or "-",
            patient.dob.strftime('%d.%m.%Y') if patient.dob else "-",
            gender_display,
            patient.nationality or "-",
            patient.passport_number or "-",
            len(patient.encounters),
            patient.created_at.strftime('%d.%m.%Y %H:%M')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_encounters_to_excel(encounters):
    """Export encounter list to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Muayeneler"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="764ba2", end_color="764ba2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Muayene ID", "Hasta Adı", "Telefon", "Tarih", "Durum",
        "Saç Ekimi", "Diş", "Göz", "Otel", "Not", "Oluşturulma"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, encounter in enumerate(encounters, 2):
        # Check which modules are used
        has_hair = bool(encounter.hair_annotations)
        has_dental = bool(encounter.dental_procedures)
        has_eye = bool(encounter.eye_refraction or encounter.eye_treatments)
        has_hotel = bool(encounter.hotel_reservations)
        
        data = [
            encounter.id,
            encounter.patient.full_name,
            encounter.patient.phone or "-",
            encounter.date.strftime('%d.%m.%Y') if encounter.date else "-",
            encounter.status.upper(),
            "✓" if has_hair else "-",
            "✓" if has_dental else "-",
            "✓" if has_eye else "-",
            "✓" if has_hotel else "-",
            encounter.note[:50] + "..." if encounter.note and len(encounter.note) > 50 else (encounter.note or "-"),
            encounter.created_at.strftime('%d.%m.%Y %H:%M')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_leads_to_excel(leads):
    """Export lead list to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lead'ler"
    
    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="11998e", end_color="11998e", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        "Lead ID", "Ad Soyad", "Telefon", "E-posta", "İlgilenen Hizmet",
        "Kaynak", "Durum", "Hasta Oldu", "Mesaj", "Oluşturulma"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Data rows
    for row_num, lead in enumerate(leads, 2):
        # Format service name
        service_map = {
            'hair': 'Saç Ekimi',
            'dental': 'Diş',
            'eye': 'Göz',
            'hotel': 'Otel'
        }
        service = service_map.get(lead.interested_service, lead.interested_service or '-')
        
        data = [
            lead.id,
            lead.full_name,
            lead.phone or "-",
            lead.email or "-",
            service,
            lead.source.upper() if lead.source else "-",
            lead.status.upper(),
            "Evet" if lead.is_converted else "Hayır",
            lead.message[:50] + "..." if lead.message and len(lead.message) > 50 else (lead.message or "-"),
            lead.created_at.strftime('%d.%m.%Y %H:%M')
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
